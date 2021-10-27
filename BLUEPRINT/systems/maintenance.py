# bluemira is an integrated inter-disciplinary design tool for future fusion
# reactors. It incorporates several modules, some of which rely on other
# codes, to carry out a range of typical conceptual fusion reactor design
# activities.
#
# Copyright (C) 2021 M. Coleman, J. Cook, F. Franza, I.A. Maione, S. McIntosh, J. Morris,
#                    D. Short
#
# bluemira is free software; you can redistribute it and/or
# modify it under the terms of the GNU Lesser General Public
# License as published by the Free Software Foundation; either
# version 2.1 of the License, or (at your option) any later version.
#
# bluemira is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# Lesser General Public License for more details.
#
# You should have received a copy of the GNU Lesser General Public
# License along with bluemira; if not, see <https://www.gnu.org/licenses/>.

"""
Some pretty outdated and crappy remote maintenance routines
"""
# flake8: noqa  (deprecated mess)
import numpy as np
import pandas as pd
from pandas import DataFrame
import itertools
import sys
import getpass
from typing import Type

from bluemira.base.look_and_feel import bluemira_print
from bluemira.base.constants import GRAVITY
from bluemira.base.parameter import ParameterFrame

from BLUEPRINT.base.file import get_BP_path
from BLUEPRINT.systems.baseclass import ReactorSystem


if sys.platform != "darwin" and getpass.getuser() != "mcintos":  # !!!
    if sys.platform == "windows":
        import win32com.client
        from openpyxl import Workbook


class RMMetrics(ReactorSystem):
    config: Type[ParameterFrame]
    inputs: dict

    default_params = [["n_TF", "Number of TF coils", 16, "N/A", None, "Input"]]
    normalise = True
    ref = {"RMTFI": 322.283, "RMSI": 554.095}

    def __init__(self, config, inputs):
        self.config = config
        self.inputs = inputs

        self._init_params(self.config)

        self.segments = self.inputs["BBgeometry"]
        self.get_AHPw()
        self.metrics()
        self.calc_RMTFI()
        self.calc_RMSI()
        self.calc_RM_duration()
        if self.normalise is True:
            self.normalise(self.ref)

    def get_AHPw(self):
        surdir = get_BP_path("Data/RemoteMaintenance")
        file = "RMTFI_Comparison_Rev_I.xlsx"
        fp = surdir + "/" + file

        AHP = pd.read_excel(fp, sheetname="AHP")
        RMTFIw = AHP[AHP.index < 8]
        RMSIw = AHP[AHP.index > 8]
        RMSIw.rename(
            index=str,
            columns={"TECHNICAL Pairwise " "matrix": "SPEED Pairwise matrix"},
            inplace=True,
        )
        self.AHPw = {}
        a = RMTFIw.set_index(RMTFIw["TECHNICAL Pairwise matrix"])
        b = RMSIw.set_index(RMSIw["SPEED Pairwise matrix"])
        self.AHPw["RMTFI"] = a["Summing to 1"].dropna().to_dict()
        self.AHPw["RMSI"] = b["Summing to 1"].dropna().to_dict()

    def metrics(self):
        self.col = ["Raw metric", "RMTFI weight", "RMSI weight"]
        m = DataFrame(columns=self.col)
        for x in self.AHPw["RMTFI"].keys():
            if "blanket segments" in x:
                m = pd.concat([m, self.seg_num()])
            elif "Kinematic" in x:
                m = pd.concat([m, self.kinematic_steps()])
            else:
                m = pd.concat([m, self.submetrics(x)])
        m.index = pd.MultiIndex.from_tuples(m.index, names=["A", "B"])
        m["RMTFI weighted"] = m["Raw metric"] * m["RMTFI weight"]
        m["RMSI weighted"] = m["Raw metric"] * m["RMSI weight"]
        self.m = m

    def calc_RM_duration(self):
        r = RMDB()
        d = r.get_values(self.params.n_TF)
        self.FM = d["Full maintenance"]
        self.DM = d["Divertor maintenance"]
        return

    def calc_RMTFI(self):
        r = []
        for metric in self.m.index.levels[0]:
            r.append(max(self.m.loc[metric]["RMTFI weighted"]))
        self.RMTFI = sum(r)
        return

    def calc_RMSI(self):
        na = "N/A"
        r = list(self.m.query("B==@na")["RMSI weighted"].values) * 4  # 1 below
        for segment in self.m.index.levels[1]:
            r.append(sum(self.m.query("B==@segment")["RMSI weighted"]))
        self.RMSI = sum(r)
        return

    def seg_num(self):
        weightT = self.AHPw["RMTFI"]["Total number of blanket segments"]
        weightS = self.AHPw["RMSI"]["Total number of blanket segments"]
        data = np.array([self.inputs["n_BB_seg"], weightT, weightS]).reshape(1, 3)
        return DataFrame(
            data, index=[("Total number of blanket segments", "N/A")], columns=self.col
        )

    def submetrics(self, name):
        metric, weightT, weightS = [], [], []
        index1, index2 = [], []
        for key in self.segments.keys():
            index1.append(name)
            index2.append(key)
            if "slenderness" in name:
                metric.append(self.manifold_slenderness(self.segments[key]))
                weightT.append(self.AHPw["RMTFI"]["Manifold slenderness ratio"])
                weightS.append(self.AHPw["RMSI"]["Manifold slenderness ratio"])
            elif "Tooling" in name:
                metric.append(self.equivalent_tooling(self.segments[key]))
                weightT.append(self.AHPw["RMTFI"]["Equivalent Tooling Stress"])
                weightS.append(self.AHPw["RMSI"]["Equivalent Tooling Stress"])
            elif "Interface" in name:
                metric.append(self.equivalent_interface(self.segments[key]))
                weightT.append(self.AHPw["RMTFI"]["Equivalent Interface Stress"])
                weightS.append(self.AHPw["RMSI"]["Equivalent Interface Stress"])
            elif "Torque" in name:
                metric.append(self.torque(self.segments[key]))
                weightT.append(self.AHPw["RMTFI"]["Torque"])
                weightS.append(self.AHPw["RMSI"]["Torque"])
            elif "Area ratio" in name:
                metric.append(self.area_ratio(self.segments[key]))
                weightT.append(self.AHPw["RMTFI"]["Area ratio"])
                weightS.append(self.AHPw["RMSI"]["Area ratio"])
        arrays = [np.array(index1), np.array(index2)]
        data = np.array([metric, weightT, weightS]).T
        return DataFrame(data, index=arrays, columns=self.col)

    def kinematic_steps(self):
        kine, weightT, weightS = [], [], []
        index1, index2 = [], []
        if self.inputs["up_shift"]:
            for segment in self.segments.keys():
                index1.append("Number of kinematic steps")
                index2.append(segment)
                if "LIBS" in segment:
                    kine.append(4)
                elif "RIBS" in segment:
                    kine.append(2)
                elif "COBS" in segment:
                    kine.append(1)
                else:
                    kine.append(2)
                weightT.append(self.AHPw["RMTFI"]["Kinematic steps to remove BS"])
                weightS.append(self.AHPw["RMSI"]["Kinematic steps to remove BS"])
        else:
            for segment in self.segments.keys():
                index1.append("Number of kinematic steps")
                index2.append(segment)
                if "LIBS" in segment:
                    kine.append(4)
                elif "RIBS" in segment:
                    kine.append(2)
                elif "COBS" in segment:
                    kine.append(3)
                else:
                    kine.append(4)
                weightT.append(self.AHPw["RMTFI"]["Kinematic steps to remove BS"])
                weightS.append(self.AHPw["RMSI"]["Kinematic steps to remove BS"])
        arrays = [np.array(index1), np.array(index2)]
        data = np.array([kine, weightT, weightS]).T
        return DataFrame(data, index=arrays, columns=self.col)

    def manifold_slenderness(self, segment):
        return segment["Length"] / segment["Radius of gyration"]

    def equivalent_tooling(self, segment):
        return (
            segment["Volume"]
            * segment["Density"]
            * GRAVITY
            / self.inputs["VV_A_enclosed"]
            / 1000
        )

    def equivalent_interface(self, segment):
        return (
            segment["Volume"]
            * segment["Density"]
            * GRAVITY
            / segment["Visible CSA"]
            / 1000
        )

    def torque(self, segment):
        cog_x = segment["Centre of gravity"]["x"]
        cog_y = segment["Centre of gravity"]["y"]
        lift_x = segment["Lift point"]["x"]
        lift_y = segment["Lift point"]["y"]
        cog_offset = ((cog_x - lift_x) ** 2 + (cog_y - lift_y) ** 2) ** 0.5
        return segment["Volume"] * segment["Density"] * GRAVITY * cog_offset / 1000

    def area_ratio(self, segment):
        return segment["Visible CSA"] / self.inputs["VV_A_enclosed"]

    def normalise(self, ref):
        self.RMTFI = ref["RMTFI"] / self.RMTFI
        self.RMSI = ref["RMSI"] / self.RMSI


class RMDB:
    def __init__(self):
        datadir = get_BP_path("Data/RemoteMaintenance")
        file = "RMDBdata"
        self.filename = datadir + "/" + file + ".json"
        self.load_RMDB()

    def load_RMDB(self):
        self.RMDB = pd.read_json(self.filename, orient="columns")

    def get_values(
        self,
        n_TF,
        type_="Hot Cell",
        ob_blankets_per_segmt=3,
        ib_blankets_per_segmt=2,
        cassettes_per_segmt=3,
        coolant_pipes_per_blanket=4,
        lipb_pipes_per_blanket=2,
        drain_pipes_per_blanket=2,
        coolant_pipes_per_cassette=2,
        n_systems=4,
    ):
        loc = self.RMDB.loc[
            (self.RMDB["n_TF"] == n_TF)
            & (self.RMDB["Type"] == type_)
            & (self.RMDB["OB_blankets_per_segmt"] == ob_blankets_per_segmt)
            & (self.RMDB["IB_blankets_per_segmt"] == ib_blankets_per_segmt)
            & (self.RMDB["Cassettes_per_segmt"] == cassettes_per_segmt)
            & (self.RMDB["Coolant_pipes_per_blanket"] == coolant_pipes_per_blanket)
            & (self.RMDB["LiPb_pipes_per_blanket"] == lipb_pipes_per_blanket)
            & (self.RMDB["Drain_pipes_per_blanket"] == drain_pipes_per_blanket)
            & (self.RMDB["Coolant_pipes_per_cassette"] == coolant_pipes_per_cassette)
            & (self.RMDB["n_systems"] == n_systems)
        ]
        full = loc["Full RM duration (days)"].values[0]
        div = loc["Divertor RM duration (days)"].values[0]
        d = {"Full maintenance": full, "Divertor maintenance": div}
        return d


class BuildRMDB:
    def __init__(
        self,
        varmin=[14, 3, 2, 3, 2, 1, 1, 2],
        varmax=[20, 4, 2, 5, 4, 2, 2, 4],
        build=False,
    ):
        self.labels = [
            "n_TF",
            "OB_blankets_per_segmt",
            "IB_blankets_per_segmt",
            "Cassettes_per_segmt",
            "Coolant_pipes_per_blanket",
            "LiPb_pipes_per_blanket",
            "Drain_pipes_per_blanket",
            "Coolant_pipes_per_cassette",
        ]
        self.varmin = varmin
        self.varmax = varmax
        if build is True:
            self.build()
            self.write("RMDBdata")

    def build(self):
        self.RMDB = DataFrame()
        kr = [0] * len(self.varmin)
        for i in range(len(self.varmin)):
            kr[i] = np.arange(self.varmin[i], self.varmax[i] + 1)
        for RDI in itertools.product(*kr):
            print(RDI)
            dp = self.data_point(list(RDI))
            self.store_dp(dp)
        return

    def write(self, fname):
        datadir = get_BP_path("Data")
        self.filename = datadir + "/" + fname + ".json"
        bluemira_print("Writing {0}".format(self.filename))
        self.RMDB.to_json(self.filename, orient="columns")

    def input_RMdurations(self, R_D_Inputs):
        datadir = "XXXXXXXXXXXXXXXXXXXXXX"
        finput = "RDI.xlsx"
        fp1 = datadir + "\\" + finput
        wb = Workbook()
        ws = wb.create_sheet("Sheet1")
        ws["D3"] = R_D_Inputs[0]
        ws["D4"] = R_D_Inputs[1]
        ws["D5"] = R_D_Inputs[2]
        ws["D6"] = R_D_Inputs[1] + R_D_Inputs[2]
        ws["D7"] = R_D_Inputs[3]
        ws["D8"] = R_D_Inputs[4]
        ws["D9"] = R_D_Inputs[5]
        ws["D10"] = R_D_Inputs[6]
        ws["D11"] = R_D_Inputs[7]
        ws["D12"] = R_D_Inputs[4] / 2 + R_D_Inputs[5] * 2 / 3 + R_D_Inputs[6] * 2 / 3
        ws["D13"] = R_D_Inputs[4] / 2 + R_D_Inputs[6] / 3
        wb.save(fp1)
        input_df = DataFrame(R_D_Inputs, index=self.labels)
        return input_df

    def read_output(self):
        """
        Ceci t'a bcp aide: http://stackoverflow.com/questions/41784468/update-links-in-for-excel-spreadsheet-using-python
        """  # noqa (W505)
        application = win32com.client.Dispatch("Excel.Application")
        application.Visible = False
        application.DisplayAlerts = False
        application.AskToUpdateLinks = False
        datadir = "XXXXXXXXXXXXXXXXXXXXXXXXXXXX"
        foutput = "RM durations mc2.xlsx"
        workbook = application.Workbooks.open(datadir + "\\" + foutput)
        workbook.UpdateLink(Name=workbook.LinkSources())
        workbook.RefreshAll()
        workbook.Save()
        workbook.Close()
        del workbook
        fp2 = datadir + "\\" + foutput
        output_df = pd.read_excel(fp2, sheetname="Durations Output")
        return output_df

    def data_point(self, r_d_inputs):
        rdinput = self.input_RMdurations(r_d_inputs)
        rmoutput = self.read_output()
        rdinput = rdinput.transpose()
        rdinput = rdinput.append([rdinput] * 9, ignore_index=True)
        dp = pd.concat([rdinput, rmoutput], axis=1)
        return dp

    def store_dp(self, dp):
        self.RMDB = pd.concat([self.RMDB, dp], ignore_index=True)
        return


if __name__ == "__main__":
    from BLUEPRINT import test

    test()
